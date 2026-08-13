# app/services/export_service.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from collections import defaultdict

from app.models.convocatoria import Convocatoria
from app.models.convocatoria_entry import ConvocatoriaEntry


def generate_convocatoria_excel(db, convocatoria):
    entries = db.query(ConvocatoriaEntry).filter(
        ConvocatoriaEntry.convocatoria_id == convocatoria.id,
        ConvocatoriaEntry.selected == True,
    ).all()

    # Agrupa entries por nadador
    by_swimmer = defaultdict(list)
    for e in entries:
        by_swimmer[e.swimmer_id].append(e)

    max_events = convocatoria.competition.max_events_per_swimmer

    wb = Workbook()
    ws = wb.active
    ws.title = "Convocatoria"

    fixed_headers = ["Nombres", "Apellidos", "RUT", "Género", "Fecha de Nacimiento", "Comuna", "Instituto", "Teléfono", "Correo Electrónico"]
    dynamic_headers = []
    for i in range(1, max_events + 1):
        dynamic_headers += [f"N°Prueba", "Tiempo"]

    headers = fixed_headers + dynamic_headers
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    row_num = 2
    for swimmer_id, swimmer_entries in by_swimmer.items():
        swimmer = swimmer_entries[0].swimmer
        row = [
            f"{swimmer.first_name_1} {swimmer.first_name_2 or ''}".strip(),
            f"{swimmer.last_name_1} {swimmer.last_name_2 or ''}".strip(),
            swimmer.document_id or "",
            swimmer.gender.value if swimmer.gender else "",
            swimmer.birth_date.strftime("%d/%m/%Y") if swimmer.birth_date else "",
            swimmer.comuna or "",
            swimmer.institution or "",
            swimmer.phone or "",
            swimmer.email or "",
        ]

        for i in range(max_events):
            if i < len(swimmer_entries):
                e = swimmer_entries[i]
                time_display = _seconds_to_display(float(e.best_time_seconds)) if e.best_time_seconds is not None else "Sin marca"
                row += [e.event_type.name, time_display]
            else:
                row += ["", ""]

        for col, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col, value=value)
        row_num += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _seconds_to_display(seconds: float) -> str:
    minutes = int(seconds // 60)
    remaining = seconds - minutes * 60
    return f"{minutes}:{remaining:05.2f}" if minutes > 0 else f"{remaining:.2f}"