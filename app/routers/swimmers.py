# app/routers/swimmers.py
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime, date
from openpyxl import Workbook
from io import BytesIO
from pydantic import BaseModel
from openpyxl.styles import Font, PatternFill
from fastapi.responses import StreamingResponse, Response
from PIL import Image

import base64, io


from app.core.deps import get_db, get_current_user
from app.models.swimmer import Swimmer, SwimmerStatus
from app.models.time_record import TimeRecord, TimeSource
from app.models.time_split import TimeSplit
from app.schemas.time_record import TimeRecordCreate, TimeRecordUpdate
from app.models.swimmer_metric import SwimmerMetric
from app.models.convocatoria_entry import ConvocatoriaEntry
from app.schemas.swimmer import SwimmerCreate, SwimmerUpdate, SwimmerStatusUpdate, SwimmerOut, SwimmerListOut
from app.models.gym_record import GymRecord
from app.schemas.export import RosterExportRequest
from app.utils.rut_validator import validate_rut, normalize_rut
from app.utils.rut_auth import rut_default_password
from app.core.security import hash_password

router = APIRouter(prefix="/swimmers", tags=["swimmers"], dependencies=[Depends(get_current_user)])

SPLIT_TOLERANCE = 0.05  # segundos de margen entre suma de parciales y tiempo total

FIELD_LABELS = {
    "first_name_1": "Primer Nombre", "first_name_2": "Segundo Nombre",
    "last_name_1": "Primer Apellido", "last_name_2": "Segundo Apellido",
    "document_id": "RUT", "birth_date": "Fecha de Nacimiento", "gender": "Género",
    "category": "Categoría", "comuna": "Comuna", "institution": "Institución",
    "phone": "Teléfono", "email": "Correo Electrónico", "profile": "Perfil",
    "is_federated": "Federado", "status": "Estado",
}

class PhotoUpload(BaseModel):
    photo_base64: str


def _validate_and_build_splits(splits_in, time_seconds: float):
    if not splits_in:
        return None
    cumulative = 0.0
    built = []
    for s in splits_in:
        cumulative += s.segment_seconds
        built.append(TimeSplit(
            distance_mark=s.distance_mark,
            segment_seconds=s.segment_seconds,
            cumulative_seconds=round(cumulative, 2),
        ))
    if abs(cumulative - time_seconds) > SPLIT_TOLERANCE:
        raise HTTPException(
            status_code=400,
            detail=f"La suma de los parciales ({cumulative:.2f}s) no coincide con el tiempo total ({time_seconds:.2f}s).",
        )
    return built


def _serialize_time_record(r: TimeRecord) -> dict:
    """Serialización única y consistente de un TimeRecord, usada por todos los endpoints de lectura."""
    return {
        "id": r.id,
        "event_type_id": r.event_type_id,
        "event_name": r.event_type.name,
        "time_seconds": float(r.time_seconds),
        "recorded_date": r.recorded_date.isoformat(),
        "pool_length": r.pool_length,
        "location_note": r.location_note,
        "competition_name": r.competition.name if r.competition else None,
        "split_increment": r.split_increment,
        "splits": [
            {
                "distance_mark": s.distance_mark,
                "segment_seconds": float(s.segment_seconds),
                "cumulative_seconds": float(s.cumulative_seconds),
            }
            for s in r.splits
        ],
    }


# ──────────────────────────────────────────────────────────────
# Nadadores — CRUD
# ──────────────────────────────────────────────────────────────

@router.get("", response_model=list[SwimmerListOut])
def list_swimmers(
    status: Optional[str] = Query(None), category: Optional[str] = None,
    profile: Optional[str] = None, is_federated: Optional[bool] = None,
    search: Optional[str] = None, db: Session = Depends(get_db),
):
    query = db.query(Swimmer)
    if status: query = query.filter(Swimmer.status == status)
    if category: query = query.filter(Swimmer.category == category)
    if profile: query = query.filter(Swimmer.profile == profile)
    if is_federated is not None: query = query.filter(Swimmer.is_federated == is_federated)
    if search:
        like = f"%{search}%"
        query = query.filter(
            (Swimmer.first_name_1.ilike(like)) | (Swimmer.last_name_1.ilike(like)) |
            (Swimmer.first_name_2.ilike(like)) | (Swimmer.last_name_2.ilike(like))
        )

    swimmers = query.order_by(Swimmer.last_name_1).all()

    return [
        SwimmerListOut(
            **{k: v for k, v in s.__dict__.items() if k != "_sa_instance_state"},
            has_photo=s.photo_base64 is not None,
        )
        for s in swimmers
    ]


@router.get("/{swimmer_id}", response_model=SwimmerOut)
def get_swimmer(swimmer_id: int, db: Session = Depends(get_db)):
    swimmer = db.query(Swimmer).filter(Swimmer.id == swimmer_id).first()
    if not swimmer:
        raise HTTPException(status_code=404, detail="Nadador no encontrado")
    return swimmer


@router.post("", response_model=SwimmerOut, status_code=201)
def create_swimmer(payload: SwimmerCreate, db: Session = Depends(get_db)):
    if payload.document_id and not validate_rut(payload.document_id):
        raise HTTPException(status_code=400, detail="RUT inválido")

    data = payload.model_dump()
    if data.get("document_id"):
        data["document_id"] = normalize_rut(data["document_id"])

        existing = db.query(Swimmer).filter(Swimmer.document_id == data["document_id"]).first()
        if existing:
            if existing.status == SwimmerStatus.DELETED:
                raise HTTPException(
                    status_code=409,
                    detail=f"Este RUT pertenece a un nadador eliminado ({existing.first_name_1} {existing.last_name_1}).",
                )
            raise HTTPException(status_code=400, detail="Este RUT ya existe en el sistema")

    swimmer = Swimmer(**data, status=SwimmerStatus.ACTIVE)
    if swimmer.birth_date:
        swimmer.category = swimmer.compute_category()

    db.add(swimmer)
    db.commit()
    db.refresh(swimmer)
    return swimmer


@router.patch("/{swimmer_id}", response_model=SwimmerOut)
def update_swimmer(swimmer_id: int, payload: SwimmerUpdate, db: Session = Depends(get_db)):
    swimmer = db.query(Swimmer).filter(Swimmer.id == swimmer_id).first()
    if not swimmer:
        raise HTTPException(status_code=404, detail="Nadador no encontrado")

    data = payload.model_dump(exclude_unset=True)

    if "document_id" in data and data["document_id"]:
        if not validate_rut(data["document_id"]):
            raise HTTPException(status_code=400, detail="RUT inválido")
        data["document_id"] = normalize_rut(data["document_id"])

        other = db.query(Swimmer).filter(
            Swimmer.document_id == data["document_id"],
            Swimmer.id != swimmer_id,
        ).first()
        if other:
            detail = (
                f"Este RUT pertenece a un nadador eliminado ({other.first_name_1} {other.last_name_1})."
                if other.status == SwimmerStatus.DELETED
                else "Este RUT ya existe en el sistema"
            )
            raise HTTPException(status_code=409 if other.status == SwimmerStatus.DELETED else 400, detail=detail)

    for field, value in data.items():
        setattr(swimmer, field, value)

    if "birth_date" in data:
        swimmer.category = swimmer.compute_category()

    db.add(swimmer)
    db.commit()
    db.refresh(swimmer)
    return swimmer


@router.patch("/{swimmer_id}/status", response_model=SwimmerOut)
def update_swimmer_status(swimmer_id: int, payload: SwimmerStatusUpdate, db: Session = Depends(get_db)):
    swimmer = db.query(Swimmer).filter(Swimmer.id == swimmer_id).first()
    if not swimmer:
        raise HTTPException(status_code=404, detail="Nadador no encontrado")

    swimmer.status = payload.status
    swimmer.status_reason = payload.reason
    swimmer.status_updated_at = datetime.now()

    db.add(swimmer)
    db.commit()
    db.refresh(swimmer)
    return swimmer


@router.delete("/{swimmer_id}", status_code=204)
def hard_delete_swimmer(swimmer_id: int, db: Session = Depends(get_db)):
    swimmer = db.query(Swimmer).filter(Swimmer.id == swimmer_id).first()
    if not swimmer:
        raise HTTPException(status_code=404, detail="Nadador no encontrado")

    try:
        # Limpia registros relacionados que NO tienen cascade definido en el modelo.
        # (metrics, time_records y attendances sí tienen cascade="all, delete-orphan"
        # en la relationship de Swimmer, así que esos se borran solos al hacer db.delete)
        db.query(GymRecord).filter(GymRecord.swimmer_id == swimmer_id).delete()
        db.query(ConvocatoriaEntry).filter(ConvocatoriaEntry.swimmer_id == swimmer_id).delete()

        db.delete(swimmer)
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="No se pudo eliminar: el nadador tiene registros asociados que impiden el borrado."
        )


@router.post("/export/custom")
def export_custom_roster(payload: RosterExportRequest, db: Session = Depends(get_db)):

    query = db.query(Swimmer).filter(Swimmer.status != SwimmerStatus.DELETED if not payload.status else Swimmer.status == payload.status)
    if payload.category:
        query = query.filter(Swimmer.category == payload.category)
    if payload.profile:
        query = query.filter(Swimmer.profile == payload.profile)
    if payload.is_federated is not None:
        query = query.filter(Swimmer.is_federated == payload.is_federated)

    swimmers = query.order_by(Swimmer.last_name_1).all()
    fields = payload.fields or list(FIELD_LABELS.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Nadadores"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col, f in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col, value=FIELD_LABELS.get(f, f))
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    def get_value(s, field):
        if field == "birth_date":
            return s.birth_date.strftime("%d/%m/%Y") if s.birth_date else ""
        if field == "gender":
            return s.gender.value if s.gender else ""
        if field == "profile":
            return {"COMPETITIVE": "Competitivo", "FORMATIVE": "Formativo"}.get(s.profile.value if s.profile else None, "")
        if field == "is_federated":
            return "Sí" if s.is_federated else "No"
        if field == "status":
            return {"ACTIVE": "Activo", "FROZEN": "Congelado"}.get(s.status.value, s.status.value)
        return getattr(s, field, "") or ""

    for row_num, s in enumerate(swimmers, start=2):
        for col, f in enumerate(fields, start=1):
            ws.cell(row=row_num, column=col, value=get_value(s, f))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="nadadores_export.xlsx"'}
    )


@router.get("/export/fields")
def get_export_fields():
    return [{"value": k, "label": v} for k, v in FIELD_LABELS.items()]


# ──────────────────────────────────────────────────────────────
# Métricas físicas (peso / altura / envergadura)
# ──────────────────────────────────────────────────────────────

@router.get("/{swimmer_id}/metrics")
def get_swimmer_metrics(swimmer_id: int, db: Session = Depends(get_db)):
    swimmer = db.query(Swimmer).filter(Swimmer.id == swimmer_id).first()
    if not swimmer:
        raise HTTPException(status_code=404, detail="Nadador no encontrado")

    return db.query(SwimmerMetric).filter(
        SwimmerMetric.swimmer_id == swimmer_id
    ).order_by(SwimmerMetric.recorded_at.desc()).all()


@router.post("/{swimmer_id}/metrics")
def add_swimmer_metric(
    swimmer_id: int,
    weight_kg: Optional[float] = None,
    height_cm: Optional[float] = None,
    wingspan_cm: Optional[float] = None,
    notes: Optional[str] = None,
    db: Session = Depends(get_db),
):
    swimmer = db.query(Swimmer).filter(Swimmer.id == swimmer_id).first()
    if not swimmer:
        raise HTTPException(status_code=404, detail="Nadador no encontrado")

    metric = SwimmerMetric(
        swimmer_id=swimmer_id, recorded_at=date.today(),
        weight_kg=weight_kg, height_cm=height_cm, wingspan_cm=wingspan_cm, notes=notes,
    )
    db.add(metric)
    db.commit()
    db.refresh(metric)
    return metric


# ──────────────────────────────────────────────────────────────
# Gimnasio / Fuerza
# ──────────────────────────────────────────────────────────────

@router.get("/{swimmer_id}/gym")
def get_gym_records(swimmer_id: int, db: Session = Depends(get_db)):
    """Devuelve el RM más reciente por ejercicio (para la vista resumen)."""
    records = db.query(GymRecord).filter(GymRecord.swimmer_id == swimmer_id).order_by(GymRecord.recorded_at.desc()).all()

    latest_by_exercise = {}
    for r in records:
        if r.exercise_id not in latest_by_exercise:
            latest_by_exercise[r.exercise_id] = r

    return [
        {"exercise_id": r.exercise_id, "exercise_name": r.exercise.name, "one_rm_kg": float(r.one_rm_kg)}
        for r in latest_by_exercise.values()
    ]


@router.get("/{swimmer_id}/gym/{exercise_id}/history")
def get_gym_history(swimmer_id: int, exercise_id: int, db: Session = Depends(get_db)):
    records = db.query(GymRecord).filter(
        GymRecord.swimmer_id == swimmer_id, GymRecord.exercise_id == exercise_id
    ).order_by(GymRecord.recorded_at.desc()).all()

    return [
        {"id": r.id, "one_rm_kg": float(r.one_rm_kg), "recorded_at": r.recorded_at.isoformat()}
        for r in records
    ]


@router.post("/{swimmer_id}/gym/{exercise_id}")
def add_gym_record(swimmer_id: int, exercise_id: int, one_rm_kg: float, db: Session = Depends(get_db)):
    """Agrega un registro NUEVO al historial (no sobreescribe)."""
    record = GymRecord(swimmer_id=swimmer_id, exercise_id=exercise_id, one_rm_kg=one_rm_kg)
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.delete("/{swimmer_id}/gym/record/{record_id}", status_code=204)
def delete_gym_record(swimmer_id: int, record_id: int, db: Session = Depends(get_db)):
    db.query(GymRecord).filter(
        GymRecord.id == record_id, GymRecord.swimmer_id == swimmer_id
    ).delete()
    db.commit()


@router.delete("/{swimmer_id}/gym/{exercise_id}", status_code=204)
def delete_swimmer_gym_history(swimmer_id: int, exercise_id: int, db: Session = Depends(get_db)):
    """Elimina TODO el historial de este nadador para este ejercicio (no borra el ejercicio en sí)."""
    db.query(GymRecord).filter(
        GymRecord.swimmer_id == swimmer_id, GymRecord.exercise_id == exercise_id
    ).delete()
    db.commit()


# ──────────────────────────────────────────────────────────────
# Tiempos / Pruebas / Parciales
# ──────────────────────────────────────────────────────────────

@router.get("/{swimmer_id}/times/grouped")
def get_times_grouped(swimmer_id: int, db: Session = Depends(get_db)):
    """Nivel 1: solo los nombres (y distancia) de las pruebas que el nadador tiene registradas."""
    records = db.query(TimeRecord).filter(TimeRecord.swimmer_id == swimmer_id).all()
    seen = {}
    for r in records:
        seen[r.event_type_id] = {"event_name": r.event_type.name, "distance_m": r.event_type.distance_m,}
    return [{"event_type_id": k, "event_name": v["event_name"], "distance_m": v["distance_m"]} for k, v in seen.items()]


@router.get("/{swimmer_id}/times")
def get_swimmer_times(
    swimmer_id: int,
    event_type_id: Optional[int] = None,
    sort: str = Query("date_desc"),
    db: Session = Depends(get_db),
):
    swimmer = db.query(Swimmer).filter(Swimmer.id == swimmer_id).first()
    if not swimmer:
        raise HTTPException(status_code=404, detail="Nadador no encontrado")

    query = db.query(TimeRecord).filter(TimeRecord.swimmer_id == swimmer_id)
    if event_type_id:
        query = query.filter(TimeRecord.event_type_id == event_type_id)

    order_map = {
        "date_desc": TimeRecord.recorded_date.desc(),
        "date_asc": TimeRecord.recorded_date.asc(),
        "time_asc": TimeRecord.time_seconds.asc(),
        "time_desc": TimeRecord.time_seconds.desc(),
    }
    records = query.order_by(order_map.get(sort, TimeRecord.recorded_date.desc())).all()

    return [_serialize_time_record(r) for r in records]


@router.get("/{swimmer_id}/times/{time_id}/splits")
def get_time_splits(swimmer_id: int, time_id: int, db: Session = Depends(get_db)):
    record = db.query(TimeRecord).filter(
        TimeRecord.id == time_id, TimeRecord.swimmer_id == swimmer_id
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    return {
        "split_increment": record.split_increment,
        "splits": [
            {"distance_mark": s.distance_mark, "segment_seconds": float(s.segment_seconds), "cumulative_seconds": float(s.cumulative_seconds)}
            for s in sorted(record.splits, key=lambda x: x.distance_mark)
        ],
    }


@router.post("/{swimmer_id}/times")
def create_time_record(swimmer_id: int, payload: TimeRecordCreate, db: Session = Depends(get_db)):
    splits = _validate_and_build_splits(payload.splits, payload.time_seconds)

    record = TimeRecord(
        swimmer_id=swimmer_id, event_type_id=payload.event_type_id, time_seconds=payload.time_seconds,
        recorded_date=payload.recorded_date, pool_length=payload.pool_length,
        location_note=payload.location_note, source=TimeSource.TRAINING,
        split_increment=payload.split_increment,
    )
    if splits:
        record.splits = splits

    db.add(record)
    db.commit()
    db.refresh(record)
    return _serialize_time_record(record)


@router.patch("/{swimmer_id}/times/{time_id}")
def update_time_record(swimmer_id: int, time_id: int, payload: TimeRecordUpdate, db: Session = Depends(get_db)):
    record = db.query(TimeRecord).filter(
        TimeRecord.id == time_id, TimeRecord.swimmer_id == swimmer_id
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    data = payload.model_dump(exclude_unset=True, exclude={"splits"})
    for field, value in data.items():
        setattr(record, field, value)

    if payload.splits is not None:
        effective_total = payload.time_seconds if payload.time_seconds is not None else float(record.time_seconds)
        splits = _validate_and_build_splits(payload.splits, effective_total)
        record.splits = splits or []
        record.split_increment = payload.split_increment

    db.add(record)
    db.commit()
    db.refresh(record)
    return _serialize_time_record(record)


@router.delete("/{swimmer_id}/times/{time_id}", status_code=204)
def delete_time_record(swimmer_id: int, time_id: int, db: Session = Depends(get_db)):
    record = db.query(TimeRecord).filter(
        TimeRecord.id == time_id, TimeRecord.swimmer_id == swimmer_id
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    db.delete(record)
    db.commit()


@router.delete("/{swimmer_id}/times/event/{event_type_id}", status_code=204)
def delete_all_times_for_event(swimmer_id: int, event_type_id: int, db: Session = Depends(get_db)):
    db.query(TimeRecord).filter(
        TimeRecord.swimmer_id == swimmer_id,
        TimeRecord.event_type_id == event_type_id,
    ).delete()
    db.commit()





@router.get("/{swimmer_id}/photo")
def get_swimmer_photo(swimmer_id: int, db: Session = Depends(get_db)):
    swimmer = db.query(Swimmer).filter(Swimmer.id == swimmer_id).first()
    
    if not swimmer or not swimmer.photo_base64:
        raise HTTPException(status_code=404)
    
    encoded = swimmer.photo_base64
    
    # Quitar prefijo si existe (por compatibilidad con fotos viejas)
    if "," in encoded:
        _, encoded = encoded.split(",", 1)
        
    # Limpiar espacios y aplicar fórmula para restaurar el padding (=) si falta
    encoded = encoded.strip()
    encoded += "=" * ((4 - len(encoded) % 4) % 4)

    try:
        image_bytes = base64.b64decode(encoded)
    except Exception as e:
        print(f"Error decodificando imagen {swimmer_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Error interno al procesar la imagen")

    return Response(
        content=image_bytes,
        media_type="image/jpeg",
        headers={"Cache-Control": "public, max-age=86400"},
    )

@router.put("/{swimmer_id}/photo")
def upload_swimmer_photo(swimmer_id: int, payload: PhotoUpload, db: Session = Depends(get_db)):
    swimmer = db.query(Swimmer).filter(Swimmer.id == swimmer_id).first()
    if not swimmer:
        raise HTTPException(status_code=404, detail="Nadador no encontrado")

    header, encoded = payload.photo_base64.split(",", 1) if "," in payload.photo_base64 else ("", payload.photo_base64)
    img_bytes = base64.b64decode(encoded)
    img = Image.open(io.BytesIO(img_bytes)).convert("RGB")

    # 400px de lado máximo — de sobra para un avatar de perfil, reduce drásticamente el peso
    img.thumbnail((400, 400))
    buffer = io.BytesIO()
    img.save(buffer, format="JPEG", quality=60, optimize=True)

    swimmer.photo_base64 = "data:image/jpeg;base64," + base64.b64encode(buffer.getvalue()).decode()
    db.add(swimmer)
    db.commit()
    return {"ok": True}


@router.delete("/{swimmer_id}/photo")
def delete_swimmer_photo(swimmer_id: int, db: Session = Depends(get_db)):
    swimmer = db.query(Swimmer).filter(Swimmer.id == swimmer_id).first()
    if not swimmer:
        raise HTTPException(status_code=404, detail="Nadador no encontrado")
    swimmer.photo_base64 = None
    db.add(swimmer)
    db.commit()
    return {"ok": True}




#----------------------------------------------------------------------



@router.post("/{swimmer_id}/reset-password")
def reset_swimmer_password(swimmer_id: int, db: Session = Depends(get_db)):
    swimmer = db.query(Swimmer).filter(Swimmer.id == swimmer_id).first()
    if not swimmer:
        raise HTTPException(status_code=404, detail="Nadador no encontrado")
    if not swimmer.document_id:
        raise HTTPException(status_code=400, detail="Este nadador no tiene RUT registrado")

    default_password = rut_default_password(swimmer.document_id)
    swimmer.hashed_password = hash_password(default_password)
    swimmer.must_change_password = True
    db.add(swimmer)
    db.commit()
    return {"ok": True, "default_password": default_password}


@router.patch("/{swimmer_id}/payment-status")
def set_payment_status(swimmer_id: int, payload: dict, db: Session = Depends(get_db)):
    swimmer = db.query(Swimmer).filter(Swimmer.id == swimmer_id).first()
    if not swimmer:
        raise HTTPException(status_code=404, detail="Nadador no encontrado")
    swimmer.payment_active = payload["payment_active"]
    db.add(swimmer)
    db.commit()
    return {"ok": True}


